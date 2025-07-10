#!/usr/bin/env python3
# scripts/optimize_hma_all_stocks.py

import os
import sys

import backtrader as bt
import pandas as pd
from openpyxl import Workbook, load_workbook

# ─── PROJECT ROOT ON PATH ──────────────────────────────────────────────────────
_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from data.load_candles import load_candles
from strategies.HmaTrendStrategy import HmaTrendStrategy

# ─── USER CONFIGURATION ────────────────────────────────────────────────────────
STOCKS      = ["AXISBANK","HDFCBANK"]
START       = "2025-04-01"
END         = "2025-07-06"
RESULTS_DIR = os.path.join(_ROOT, "results")
os.makedirs(RESULTS_DIR, exist_ok=True)

# HMA grid: multiples of 40
FAST_VALS = list(range(40, 1001, 40))    # 40, 80, 120, ..., 1000
SLOW_VALS = list(range(40, 3001, 40))    # 40, 80, 120, ..., 3000

# Output file
OUTPUT_XLSX = os.path.join(RESULTS_DIR, "hma_all_stocks.xlsx")
SHEET_NAME  = "Results"

def append_to_xlsx(path, row, sheet=SHEET_NAME):
    """Append a row to an Excel sheet, creating file/header if needed."""
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet
        ws.append(list(row.keys()))
        ws.append(list(row.values()))
        wb.save(path)
    else:
        wb = load_workbook(path)
        ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)
        ws.append(list(row.values()))
        wb.save(path)

def load_done_set(path, sheet=SHEET_NAME):
    """Read existing results to skip already-done combos."""
    if not os.path.exists(path):
        return set()
    df = pd.read_excel(path, sheet_name=sheet)
    return set(zip(df['symbol'], df['fast'], df['slow']))

def optimize_all():
    done = load_done_set(OUTPUT_XLSX)
    total = len(STOCKS) * len(FAST_VALS) * len(SLOW_VALS)
    print(f"Running up to {total} backtests (fast < slow only)...\n")

    for symbol in STOCKS:
        df_feed = load_candles(symbol, START, END)
        for fast in FAST_VALS:
            for slow in SLOW_VALS:
                # only test when fast < slow
                if fast >= slow:
                    continue

                key = (symbol, fast, slow)
                if key in done:
                    continue

                cerebro = bt.Cerebro()
                cerebro.addanalyzer(bt.analyzers.SharpeRatio, _name="sharpe",
                                    timeframe=bt.TimeFrame.Minutes, riskfreerate=0.0)
                cerebro.addanalyzer(bt.analyzers.DrawDown,    _name="drawdown")
                cerebro.addanalyzer(bt.analyzers.TradeAnalyzer,_name="trades")

                data = bt.feeds.PandasData(
                    dataname=df_feed,
                    timeframe=bt.TimeFrame.Minutes,
                    compression=1
                )
                cerebro.adddata(data, name=symbol)

                cerebro.addstrategy(
                    HmaTrendStrategy,
                    fast=fast,
                    slow=slow,
                    atr_mult=0.0,
                    printlog=False
                )

                strat = cerebro.run()[0]
                sa = strat.analyzers
                raw_sh = sa.sharpe.get_analysis().get("sharperatio", None)
                sharpe = round(raw_sh, 4) if raw_sh is not None else float("nan")
                dd     = sa.drawdown.get_analysis().max.drawdown
                tr     = sa.trades.get_analysis()
                total_trades = tr.get("total", {}).get("closed", 0)
                won          = tr.get("won",   {}).get("total",  0)
                winpct       = (won / total_trades * 100) if total_trades else 0.0

                row = {
                    "symbol": symbol,
                    "fast":   fast,
                    "slow":   slow,
                    "sharpe": sharpe,
                    "max_dd": round(dd, 4),
                    "trades": total_trades,
                    "win%":   round(winpct, 1),
                }
                append_to_xlsx(OUTPUT_XLSX, row)
                done.add(key)

                print(f"{symbol:8s} f={fast:<4d} s={slow:<4d} → Sharpe {sharpe:.4f}, Win% {row['win%']:.1f}%")

if __name__ == "__main__":
    optimize_all()
